"""
パイプラインステージ定義

各ステージの処理ロジックを定義
"""
import logging
import zipfile
from pathlib import Path
from typing import Optional, Callable, List
import pandas as pd
from openpyxl import load_workbook

from config import (
    DOWNLOAD_DIR,
    RAW_DIR,
    NORMALIZED_DIR,
    PROCESSED_DIR,
    SCHEMA_DIR,
    FILENAME_TO_YEAR,
)
from src.utils.normalization import normalize_text, normalize_column_name
from src.pipeline.table_builder import process_year_data
from src.pipeline.schema_generator import process_directory_schemas

logger = logging.getLogger(__name__)


class PipelineStage:
    """パイプラインステージの基底クラス"""

    def __init__(self, name: str, description: str):
        self.name = name
        self.description = description

    def run(self, update_callback: Optional[Callable] = None, target_year: Optional[int] = None) -> bool:
        """
        ステージを実行

        Args:
            update_callback: 進捗更新用のコールバック関数
            target_year: 処理対象年度（指定しない場合は全年度）

        Returns:
            成功した場合True
        """
        raise NotImplementedError


class Stage01_ExtractToCSV(PipelineStage):
    """Stage 1: Excel/ZIP → CSV変換"""

    def __init__(self):
        super().__init__(
            name="Stage 1: Excel/ZIP to CSV",
            description="Excel/ZIPファイルをCSVに変換"
        )

    def run(self, update_callback: Optional[Callable] = None, target_year: Optional[int] = None) -> bool:
        """Excel/ZIPファイルをCSVに変換"""
        logger.info(f"Starting {self.name}")
        if target_year:
            logger.info(f"Processing only year {target_year}")

        if not DOWNLOAD_DIR.exists():
            logger.error(f"Download directory not found: {DOWNLOAD_DIR}")
            return False

        RAW_DIR.mkdir(parents=True, exist_ok=True)

        # Excel/ZIPファイルを取得
        files = list(DOWNLOAD_DIR.glob("*.xlsx")) + list(DOWNLOAD_DIR.glob("*.zip"))

        if not files:
            logger.warning(f"No Excel/ZIP files found in {DOWNLOAD_DIR}")
            return False

        total_files = len(files)

        for idx, file_path in enumerate(files, 1):
            if update_callback:
                update_callback(f"Processing {file_path.name} ({idx}/{total_files})")

            logger.info(f"Processing: {file_path.name}")

            # 年度を取得
            year = FILENAME_TO_YEAR.get(file_path.name)

            if file_path.suffix == '.zip':
                self._extract_zip_to_csv(file_path, year)
            elif file_path.suffix == '.xlsx':
                self._extract_excel_to_csv(file_path, year)

        logger.info(f"Completed {self.name}")
        return True

    def _extract_zip_to_csv(self, zip_path: Path, year: Optional[int]):
        """ZIPファイルを解凍してCSVに変換"""
        extract_dir = RAW_DIR / f"year_{year}" if year else RAW_DIR / zip_path.stem
        extract_dir.mkdir(parents=True, exist_ok=True)

        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
            logger.info(f"Extracted {zip_path.name} to {extract_dir}")

            # 解凍されたExcelファイルを処理（サブディレクトリも含む）
            for excel_file in extract_dir.rglob("*.xlsx"):
                self._extract_excel_to_csv(excel_file, year, extract_dir)

    def _extract_excel_to_csv(self, excel_path: Path, year: Optional[int], output_dir: Optional[Path] = None):
        """ExcelファイルをCSVに変換"""
        if output_dir is None:
            output_dir = RAW_DIR / f"year_{year}" if year else RAW_DIR / excel_path.stem

        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            # openpyxlでシート名を取得
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            for sheet_name in sheet_names:
                logger.info(f"  Processing sheet: {sheet_name}")

                # pandasでシートを読み込み
                df = pd.read_excel(excel_path, sheet_name=sheet_name)

                # CSVファイル名を生成
                csv_filename = f"{year}_{sheet_name}.csv" if year else f"{sheet_name}.csv"
                csv_path = output_dir / csv_filename

                # CSV保存
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                logger.info(f"    Saved: {csv_path.name}")

            wb.close()

        except Exception as e:
            logger.error(f"Error processing {excel_path.name}: {e}")


class Stage02_Normalize(PipelineStage):
    """Stage 2: テキスト正規化"""

    def __init__(self):
        super().__init__(
            name="Stage 2: Text Normalization",
            description="日本語テキストの正規化（和暦変換、記号統一等）"
        )

    def run(self, update_callback: Optional[Callable] = None, target_year: Optional[int] = None) -> bool:
        """CSVファイルのテキストを正規化"""
        logger.info(f"Starting {self.name}")

        if not RAW_DIR.exists():
            logger.error(f"Raw directory not found: {RAW_DIR}")
            return False

        NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)

        # 全CSVファイルを取得
        csv_files = list(RAW_DIR.rglob("*.csv"))

        if not csv_files:
            logger.warning(f"No CSV files found in {RAW_DIR}")
            return False

        total_files = len(csv_files)

        for idx, csv_path in enumerate(csv_files, 1):
            if update_callback:
                update_callback(f"Normalizing {csv_path.name} ({idx}/{total_files})")

            logger.info(f"Normalizing: {csv_path}")

            # 相対パスを維持して出力先を決定
            relative_path = csv_path.relative_to(RAW_DIR)
            output_path = NORMALIZED_DIR / relative_path
            output_path.parent.mkdir(parents=True, exist_ok=True)

            self._normalize_csv(csv_path, output_path)

        logger.info(f"Completed {self.name}")
        return True

    def _normalize_csv(self, input_path: Path, output_path: Path):
        """CSVファイルを正規化"""
        try:
            # CSVを読み込み
            df = pd.read_csv(input_path, encoding='utf-8-sig')

            # カラム名を正規化（normalize_text で完全な正規化を実施）
            df.columns = [normalize_text(col) for col in df.columns]

            # データを正規化（文字列型のカラムのみ）
            for col in df.columns:
                if df[col].dtype == 'object':  # 文字列型
                    df[col] = df[col].apply(
                        lambda x: normalize_text(x) if isinstance(x, str) else x
                    )

            # 正規化済みCSVを保存
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            logger.info(f"  Normalized: {output_path.name}")

        except Exception as e:
            logger.error(f"Error normalizing {input_path.name}: {e}")


class Stage03_BuildTables(PipelineStage):
    """Stage 3: RSシステム形式のテーブル構築"""

    def __init__(self):
        super().__init__(
            name="Stage 3: Build RS System Tables",
            description="正規化されたデータをRSシステム形式のテーブルに変換"
        )

    def run(self, update_callback: Optional[Callable] = None, target_year: Optional[int] = None) -> bool:
        """RSシステム形式のテーブルを構築"""
        logger.info(f"Starting {self.name}")
        if target_year:
            logger.info(f"Processing only year {target_year}")

        if not NORMALIZED_DIR.exists():
            logger.error(f"Normalized directory not found: {NORMALIZED_DIR}")
            return False

        PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

        # 年度ごとのディレクトリを処理
        year_dirs = [d for d in NORMALIZED_DIR.iterdir() if d.is_dir()]

        # 年度指定がある場合はフィルター
        if target_year:
            year_dirs = [d for d in year_dirs if d.name == f"year_{target_year}"]
            if not year_dirs:
                logger.error(f"Year directory not found: year_{target_year}")
                return False

        total_files = 0
        total_success = 0
        total_failed = 0

        for year_dir in year_dirs:
            if update_callback:
                update_callback(f"Building tables for {year_dir.name}")

            logger.info(f"Processing {year_dir.name}")

            # テーブル構築処理
            files, success, failed = process_year_data(year_dir, PROCESSED_DIR)
            total_files += files
            total_success += success
            total_failed += failed

        logger.info(
            f"Completed {self.name}: {total_files} files, "
            f"{total_success} success, {total_failed} failed"
        )
        return total_failed == 0


class Stage04_GenerateSchema(PipelineStage):
    """Stage 4: スキーマ定義の生成"""

    def __init__(self):
        super().__init__(
            name="Stage 4: Generate Schema",
            description="処理済みデータのスキーマ定義を生成"
        )

    def run(self, update_callback: Optional[Callable] = None, target_year: Optional[int] = None) -> bool:
        """スキーマ定義を生成"""
        logger.info(f"Starting {self.name}")

        if not PROCESSED_DIR.exists():
            logger.error(f"Processed directory not found: {PROCESSED_DIR}")
            return False

        SCHEMA_DIR.mkdir(parents=True, exist_ok=True)

        if update_callback:
            update_callback("Generating schema definitions")

        # スキーマ生成処理
        num_files = process_directory_schemas(
            PROCESSED_DIR, SCHEMA_DIR, max_rows=10000
        )

        logger.info(f"Completed {self.name}: {num_files} schemas generated")
        return True


# 利用可能なステージのリスト
AVAILABLE_STAGES = [
    Stage01_ExtractToCSV(),
    Stage02_Normalize(),
    Stage03_BuildTables(),
    Stage04_GenerateSchema(),
]


def get_stage_by_number(stage_num: int) -> Optional[PipelineStage]:
    """
    ステージ番号からステージを取得

    Args:
        stage_num: ステージ番号（1-4）

    Returns:
        PipelineStageオブジェクト（存在しない場合はNone）
    """
    if 1 <= stage_num <= len(AVAILABLE_STAGES):
        return AVAILABLE_STAGES[stage_num - 1]
    return None
